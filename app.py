import streamlit as st
from streamlit_drawable_canvas import st_canvas
from datetime import datetime, time
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import os

# ==================================================
# CONFIGURAÇÕES
# ==================================================
BASE_DIR = r"C:\Users\MarlonLucasRocha\OneDrive - Santa Casa de Misericórdia de Chavantes\PROJETO PYTHON"
PASTA_BASE = os.path.join(BASE_DIR, "justificativas")
PLANILHA_PATH = os.path.join(PASTA_BASE, "registros.xlsx")
LOGO_PATH = os.path.join(BASE_DIR, "imagens", "mitri_logo.png")

SETORES = [
    "Clínica Médica - PS",
    "Diarista - Neurologista",
    "Neurocirurgia",
    "UTI"
]

for setor in SETORES:
    os.makedirs(os.path.join(PASTA_BASE, setor), exist_ok=True)

st.set_page_config(page_title="Justificativa de Ponto", layout="centered")

# ==================================================
# CABEÇALHO DO FORMULÁRIO
# ==================================================
col_logo, col_text = st.columns([1, 6])
with col_logo:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, width=90)
with col_text:
    st.markdown("### **FORMULÁRIO DE JUSTIFICATIVA DO PONTO**")
    st.markdown("<span style='color:gray'>Hospital Regional Sul</span>", unsafe_allow_html=True)
st.markdown("---")

# ==================================================
# FUNÇÃO PLANILHA (BONITA)
# ==================================================
def registrar_planilha(dados):
    headers = [
        "Nome dos Médicos",
        "Horas de Plantão",
        "Data do Plantão",
        "Horário de Entrada",
        "Horário de Saída",
        "Ocupação",
        "Motivo"
    ]

    nova = not os.path.exists(PLANILHA_PATH)

    if nova:
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"

        if os.path.exists(LOGO_PATH):
            img = XLImage(LOGO_PATH)
            img.width = 160
            img.height = 70
            ws.add_image(img, "A1")

        ws.merge_cells("B2:H3")
        t = ws["B2"]
        t.value = "RELATÓRIO DE JUSTIFICATIVAS DE PLANTÃO\nHOSPITAL REGIONAL SUL"
        t.font = Font(size=14, bold=True)
        t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        header_fill = PatternFill("solid", fgColor="D9D9D9")
        border = Border(*(Side(style="thin"),)*4)

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        wb.save(PLANILHA_PATH)

    try:
        wb = load_workbook(PLANILHA_PATH)
        ws = wb.active
        row = ws.max_row + 1

        for col, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=dados[h]).border = Border(*(Side(style="thin"),)*4)

        for i in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(i)].width = 24

        wb.save(PLANILHA_PATH)

    except PermissionError:
        st.error("❌ Feche o Excel antes de gerar outro documento.")
        st.stop()

# ==================================================
# FORMULÁRIO
# ==================================================
with st.form("formulario"):
    nome = st.text_input("Nome do médico *")
    crm = st.text_input("CRM *")
    ocupacao = st.selectbox("Setor *", SETORES)
    data = st.date_input("Data do Plantão *", format="DD/MM/YYYY")

    c1, c2 = st.columns(2)
    with c1:
        hora_entrada = st.time_input("Horário de Entrada *", value=time(7, 0))
    with c2:
        hora_saida = st.time_input("Horário de Saída *", value=time(19, 0))

    motivo = st.text_area("Motivo *", height=120)
    assinatura = st.text_input("Assinatura *")

    enviar = st.form_submit_button("✅ Gerar documento")

# ==================================================
# PROCESSAMENTO (TUDO AQUI DENTRO)
# ==================================================
if enviar:
    if not nome or not crm or not motivo:
        st.error("Preencha todos os campos obrigatórios.")
        st.stop()

    # ===============================
    # FORMATOS
    # ===============================
    data_fmt = data.strftime("%d/%m/%Y")
    data_arquivo = data.strftime("%d-%m-%Y")

    hora_ent = hora_entrada.strftime("%H:%M")
    hora_sai = hora_saida.strftime("%H:%M")

    # Horas no formato HH:MM
    duracao = datetime.combine(data, hora_saida) - datetime.combine(data, hora_entrada)
    horas_formatadas = (
        f"{int(duracao.total_seconds() // 3600):02d}:"
        f"{int((duracao.total_seconds() % 3600) // 60):02d}"
    )

    # ===============================
    # CAMINHO DO PDF (AQUI NASCE pdf_path)
    # ===============================
    pasta_setor = os.path.join(PASTA_BASE, ocupacao)
    pdf_path = os.path.join(pasta_setor, f"{nome} - {data_arquivo}.pdf")

    # ===============================
    # CRIA O PDF (NADA ANTES DISSO)
    # ===============================
    c = canvas.Canvas(pdf_path, pagesize=A4)

    W, H = A4
    X = 2 * cm

    from reportlab.lib import colors
    COR_PRINCIPAL = colors.HexColor("#030D0C")
    COR_SUAVE = colors.HexColor("#E6F7F6")
    COR_ROTULO = colors.HexColor("#4D4D4D")

    # ===============================
    # CABEÇALHO (SEU, INTACTO)
    # ===============================
    if os.path.exists(LOGO_PATH):
        lw = 6 * cm
        c.drawImage(LOGO_PATH, (W - lw) / 2, H - 7 * cm, width=lw, preserveAspectRatio=True)

    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(
        W / 2,
        H - 7.0 * cm,
        "FORMULÁRIO DE JUSTIFICATIVA DO PONTO – HOSPITAL REGIONAL SUL"
    )

    c.line(X, H - 8.2 * cm, W - X, H - 8.2 * cm)

    # ===============================
    # CORPO DO PDF (BONITO)
    # ===============================
    y = H - 10 * cm
    linha = 1 * cm

    c.setStrokeColor(COR_PRINCIPAL)
    c.setLineWidth(1.6)
    c.rect(X - 0.5 * cm, y - 4.8 * cm, W - 2*X + 1 * cm, 5.6 * cm)

    def campo(y_pos, rotulo, valor):
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(COR_ROTULO)
        c.drawString(X, y_pos, rotulo)

        c.setFont("Helvetica", 10)
        c.setFillColor(colors.black)
        c.drawString(X + 6 * cm, y_pos, valor)

        return y_pos - linha

    y = campo(y, "Nome:", nome)
    y = campo(y, "CRM:", crm)
    y = campo(y, "Setor:", ocupacao)
    y = campo(y, "Data:", data_fmt)
    y = campo(y, "Horário:", f"{hora_ent} - {hora_sai}")

    # ===============================
    # JUSTIFICATIVA
    # ===============================
    y -= 0.8 * cm
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(COR_PRINCIPAL)
    c.drawString(X, y, "Justificativa:")

    c.setFillColor(COR_SUAVE)
    c.rect(X, y - 4.6 * cm, W - 2 * X, 4 * cm, fill=1, stroke=0)

    c.setStrokeColor(COR_PRINCIPAL)
    c.rect(X, y - 4.6 * cm, W - 2 * X, 4 * cm, fill=0, stroke=1)

    texto = c.beginText(X + 0.4 * cm, y - 1 * cm)
    texto.setFont("Helvetica", 10)
    texto.setLeading(14)
    texto.setFillColor(colors.black)

    for linha_txt in motivo.split("\n"):
        texto.textLine(linha_txt)

    c.drawText(texto)

    # ===============================
    # ASSINATURA (RESTAURADA)
    # ===============================
    y_ass = 6 * cm
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(COR_PRINCIPAL)
    c.drawString(X, y_ass + 1.8 * cm, "Assinatura:")

    c.setFont("Helvetica", 11)
    c.setFillColor(colors.black)
    c.drawString(X, y_ass + 1.2 * cm, assinatura)

    c.line(X, y_ass + 1.1 * cm, X + 14 * cm, y_ass + 1.1 * cm)

    # ===============================
    # RODAPÉ
    # ===============================
    c.setFont("Helvetica-Oblique", 8)
    c.setFillColor(colors.grey)
    c.drawCentredString(
        W / 2,
        1.3 * cm,
        "Documento gerado eletronicamente por meio de sistema interno da instituição."
    )

    c.save()

    # ===============================
    # PLANILHA
    # ===============================
    registrar_planilha({
        "Nome dos Médicos": nome,
        "Horas de Plantão": horas_formatadas,
        "Data do Plantão": data_fmt,
        "Horário de Entrada": hora_ent,
        "Horário de Saída": hora_sai,
        "Ocupação": ocupacao,
        "Motivo": motivo
    })

    st.success("✅ Documento gerado com sucesso!")
    st.info(f"📄 PDF salvo em: {pdf_path}")